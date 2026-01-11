#!/usr/bin/env python3
"""
azfilter.jp catalogue exporter

Pipeline:
  1) collect-pids  -> out/cache/azfilter.sqlite3 (parts from /api/part_suggestions)
  2) scrape-pages  -> cache HTML pages + extract application fitment rows
  3) export-xlsx   -> out/azfilter_parts.xlsx (Parts + Fitment sheets)

Notes:
  - The sandbox environment may not have CA certs. Use --insecure to skip TLS verification.
  - The /api/part_suggestions endpoint returns up to 10 suggestions. We split prefixes to enumerate all.
"""

from __future__ import annotations

import argparse
import dataclasses
import datetime as dt
import json
import os
import re
import sqlite3
import ssl
import sys
import time
from collections import deque
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import parse_qs, quote, urljoin, urlparse
from urllib.request import Request, urlopen

from bs4 import BeautifulSoup  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from tqdm import tqdm  # type: ignore


BASE_URL = "https://azfilter.jp"
API_PART_SUGGESTIONS = urljoin(BASE_URL, "/api/part_suggestions")
PRODUCT_URL_TMPL = urljoin(BASE_URL, "/catalogue/catalogue/{pid}")

SUGGESTIONS_LIMIT = 10

# This alphabet is used to split prefixes when /api/part_suggestions hits the limit.
# It’s intentionally conservative (letters/digits) because the API seems to match
# part numbers even if they contain spaces/hyphens.
ALPHABET = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"


def utc_now_iso() -> str:
    return dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def normalize_part_key(s: str) -> str:
    """Normalize part number to a key suitable for prefix splitting: uppercase, alnum only."""
    return re.sub(r"[^0-9A-Z]+", "", (s or "").upper())


class HttpClient:
    def __init__(self, insecure: bool, user_agent: str = "Mozilla/5.0") -> None:
        self.insecure = insecure
        self.user_agent = user_agent
        self._ssl_ctx = ssl._create_unverified_context() if insecure else ssl.create_default_context()

    def get_bytes(self, url: str, *, accept: str = "*/*") -> bytes:
        req = Request(
            url,
            headers={
                "User-Agent": self.user_agent,
                "Accept": accept,
            },
        )
        with urlopen(req, context=self._ssl_ctx, timeout=60) as r:
            return r.read()

    def get_json(self, url: str) -> Any:
        raw = self.get_bytes(url, accept="application/json, text/plain;q=0.9, */*;q=0.8")
        txt = raw.decode("utf-8", "replace")
        try:
            return json.loads(txt)
        except json.JSONDecodeError:
            # Sometimes the site returns an HTML redirect on invalid inputs.
            raise RuntimeError(f"Non-JSON response from {url}: {txt[:200]!r}")


class Db:
    def __init__(self, path: Path) -> None:
        self.path = path
        ensure_dir(path.parent)
        self.conn = sqlite3.connect(str(path))
        self.conn.row_factory = sqlite3.Row
        self._init_schema()

    def close(self) -> None:
        self.conn.close()

    def _init_schema(self) -> None:
        cur = self.conn.cursor()
        cur.executescript(
            """
            PRAGMA journal_mode=WAL;
            PRAGMA synchronous=NORMAL;

            CREATE TABLE IF NOT EXISTS parts (
              pid INTEGER PRIMARY KEY,
              part_no TEXT,
              p_brand TEXT,
              type_name TEXT,
              house_part_no TEXT,
              house_brand TEXT,
              moved_id TEXT,
              moved_part_no TEXT,
              first_seen_prefix TEXT,
              raw_json TEXT,
              created_at TEXT,
              updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS prefixes (
              prefix TEXT PRIMARY KEY,
              status TEXT NOT NULL,            -- pending|done|error
              last_count INTEGER,
              tries INTEGER NOT NULL DEFAULT 0,
              last_error TEXT,
              updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS pages (
              pid INTEGER PRIMARY KEY,
              status TEXT NOT NULL,            -- pending|done|error
              http_status INTEGER,
              last_error TEXT,
              updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS fitments (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              pid INTEGER NOT NULL,
              make TEXT,
              model TEXT,
              body TEXT,
              engine_no TEXT,
              engine_cc TEXT,
              date_from TEXT,
              date_to TEXT,
              notes TEXT,
              href TEXT,
              raw_article TEXT,
              raw_text TEXT,
              created_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_fitments_pid ON fitments(pid);
            """
        )
        self.conn.commit()

    def upsert_part_from_suggestion(self, pid: int, suggestion: Dict[str, Any], prefix: str) -> None:
        now = utc_now_iso()
        cur = self.conn.cursor()
        cur.execute(
            """
            INSERT INTO parts (
              pid, part_no, p_brand, type_name, house_part_no, house_brand,
              moved_id, moved_part_no, first_seen_prefix, raw_json, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(pid) DO UPDATE SET
              part_no=excluded.part_no,
              p_brand=excluded.p_brand,
              type_name=excluded.type_name,
              house_part_no=excluded.house_part_no,
              house_brand=excluded.house_brand,
              moved_id=excluded.moved_id,
              moved_part_no=excluded.moved_part_no,
              raw_json=excluded.raw_json,
              updated_at=excluded.updated_at
            """,
            (
                pid,
                suggestion.get("part_no"),
                suggestion.get("p_brand"),
                suggestion.get("type_name"),
                suggestion.get("house_part_no"),
                suggestion.get("house_brand"),
                str(suggestion.get("moved_id") or ""),
                suggestion.get("moved_part_no"),
                prefix,
                json.dumps(suggestion, ensure_ascii=False),
                now,
                now,
            ),
        )
        self.conn.commit()

    def mark_prefix(self, prefix: str, status: str, *, last_count: Optional[int] = None, last_error: str = "") -> None:
        now = utc_now_iso()
        cur = self.conn.cursor()
        cur.execute(
            """
            INSERT INTO prefixes(prefix, status, last_count, tries, last_error, updated_at)
            VALUES (?, ?, ?, 1, ?, ?)
            ON CONFLICT(prefix) DO UPDATE SET
              status=excluded.status,
              last_count=excluded.last_count,
              tries=prefixes.tries + 1,
              last_error=excluded.last_error,
              updated_at=excluded.updated_at
            """,
            (prefix, status, last_count, last_error, now),
        )
        self.conn.commit()

    def prefix_status(self, prefix: str) -> Optional[str]:
        cur = self.conn.cursor()
        cur.execute("SELECT status FROM prefixes WHERE prefix=?", (prefix,))
        row = cur.fetchone()
        return row["status"] if row else None

    def prefix_meta(self, prefix: str) -> Optional[sqlite3.Row]:
        cur = self.conn.cursor()
        cur.execute("SELECT prefix, status, last_count, tries, updated_at FROM prefixes WHERE prefix=?", (prefix,))
        return cur.fetchone()

    def prefix_has_children(self, prefix: str) -> bool:
        # Any longer prefix starting with this prefix counts as a child.
        cur = self.conn.cursor()
        cur.execute("SELECT 1 FROM prefixes WHERE prefix LIKE ? LIMIT 1", (prefix + "%",))
        row = cur.fetchone()
        if not row:
            return False
        # The query matches itself; ensure there's something longer.
        cur.execute("SELECT 1 FROM prefixes WHERE prefix LIKE ? AND length(prefix) > length(?) LIMIT 1", (prefix + "%", prefix))
        return cur.fetchone() is not None

    def mark_page(self, pid: int, status: str, *, http_status: Optional[int] = None, last_error: str = "") -> None:
        now = utc_now_iso()
        cur = self.conn.cursor()
        cur.execute(
            """
            INSERT INTO pages(pid, status, http_status, last_error, updated_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(pid) DO UPDATE SET
              status=excluded.status,
              http_status=excluded.http_status,
              last_error=excluded.last_error,
              updated_at=excluded.updated_at
            """,
            (pid, status, http_status, last_error, now),
        )
        self.conn.commit()

    def page_status(self, pid: int) -> Optional[str]:
        cur = self.conn.cursor()
        cur.execute("SELECT status FROM pages WHERE pid=?", (pid,))
        row = cur.fetchone()
        return row["status"] if row else None

    def list_part_pids(self) -> List[int]:
        cur = self.conn.cursor()
        cur.execute("SELECT pid FROM parts ORDER BY pid")
        return [int(r["pid"]) for r in cur.fetchall()]

    def clear_fitments_for_pid(self, pid: int) -> None:
        cur = self.conn.cursor()
        cur.execute("DELETE FROM fitments WHERE pid=?", (pid,))
        self.conn.commit()

    def insert_fitments(self, pid: int, rows: Sequence[Dict[str, Any]]) -> None:
        now = utc_now_iso()
        cur = self.conn.cursor()
        cur.executemany(
            """
            INSERT INTO fitments(
              pid, make, model, body, engine_no, engine_cc, date_from, date_to,
              notes, href, raw_article, raw_text, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    pid,
                    r.get("make"),
                    r.get("model"),
                    r.get("body"),
                    r.get("engine_no"),
                    r.get("engine_cc"),
                    r.get("date_from"),
                    r.get("date_to"),
                    r.get("notes"),
                    r.get("href"),
                    r.get("raw_article"),
                    r.get("raw_text"),
                    now,
                )
                for r in rows
            ],
        )
        self.conn.commit()

    def export_parts(self) -> List[sqlite3.Row]:
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT pid, part_no, p_brand, type_name, house_part_no, house_brand, moved_id, moved_part_no,
                   first_seen_prefix, created_at, updated_at
            FROM parts
            ORDER BY pid
            """
        )
        return cur.fetchall()

    def export_fitments(self) -> List[sqlite3.Row]:
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT pid, make, model, body, engine_no, engine_cc, date_from, date_to, notes, href, raw_article, raw_text, created_at
            FROM fitments
            ORDER BY pid, id
            """
        )
        return cur.fetchall()

    def counts(self) -> Dict[str, int]:
        cur = self.conn.cursor()
        out: Dict[str, int] = {}
        for table in ("parts", "prefixes", "pages", "fitments"):
            cur.execute(f"SELECT count(*) AS c FROM {table}")
            out[table] = int(cur.fetchone()[0])
        return out

    def prefixes_status_counts(self) -> Dict[str, int]:
        cur = self.conn.cursor()
        cur.execute("SELECT status, count(*) AS c FROM prefixes GROUP BY status")
        rows = cur.fetchall()
        out = {str(r["status"]): int(r["c"]) for r in rows}
        out.setdefault("pending", 0)
        out.setdefault("done", 0)
        out.setdefault("error", 0)
        return out

    def pages_status_counts(self) -> Dict[str, int]:
        cur = self.conn.cursor()
        cur.execute("SELECT status, count(*) AS c FROM pages GROUP BY status")
        rows = cur.fetchall()
        out = {str(r["status"]): int(r["c"]) for r in rows}
        out.setdefault("pending", 0)
        out.setdefault("done", 0)
        out.setdefault("error", 0)
        return out

    def pages_remaining_estimate(self) -> int:
        """How many part pages are not marked done yet (best effort)."""
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT count(*) AS c
            FROM parts p
            LEFT JOIN pages pg ON pg.pid = p.pid
            WHERE pg.status IS NULL OR pg.status != 'done'
            """
        )
        return int(cur.fetchone()[0])


def build_suggestions_query(prefix: str) -> str:
    # API appears to accept '*' as a wildcard and counts toward the minimal length.
    # For short prefixes, pad to length 4 with '*'.
    p = prefix.upper()
    if len(p) < 4:
        p = p + ("*" * (4 - len(p)))
    return p


def fetch_part_suggestions(client: HttpClient, prefix: str) -> List[Dict[str, Any]]:
    q = build_suggestions_query(prefix)
    url = API_PART_SUGGESTIONS + "?part_no=" + quote(q)
    data = client.get_json(url)
    if not isinstance(data, list):
        raise RuntimeError(f"Unexpected suggestions response for {prefix!r}: {type(data)}")
    return data  # list of dicts


def collect_pids(
    db: Db,
    client: HttpClient,
    *,
    sleep_s: float,
    max_prefix_len: int,
    start_prefixes: Sequence[str],
) -> None:
    queue = deque(start_prefixes)
    queued = set(start_prefixes)
    processed = 0
    last_report = time.time()

    # BFS-ish: pop from front (ok for moderate sizes)
    while queue:
        prefix = queue.popleft()
        queued.discard(prefix)

        meta = db.prefix_meta(prefix)
        if meta and meta["status"] == "done":
            # Important for resuming: if this prefix previously hit the limit and we
            # now allow deeper search, we must expand it even if it's marked done.
            last_count = meta["last_count"]
            if (
                last_count is not None
                and int(last_count) >= SUGGESTIONS_LIMIT
                and len(prefix) < max_prefix_len
                and not db.prefix_has_children(prefix)
            ):
                for ch in ALPHABET:
                    sub = prefix + ch
                    if db.prefix_status(sub) != "done" and sub not in queued:
                        queue.append(sub)
                        queued.add(sub)
            continue

        try:
            suggestions = fetch_part_suggestions(client, prefix)
            db.mark_prefix(prefix, "done", last_count=len(suggestions))

            for s in suggestions:
                pid = int(s["pid"])
                db.upsert_part_from_suggestion(pid, s, prefix)

            # If we hit the limit, split into sub-prefixes.
            if len(suggestions) >= SUGGESTIONS_LIMIT and len(prefix) < max_prefix_len:
                # Conservative split: try all alnum next characters to avoid missing items
                # that may not appear in the top-10 suggestions.
                for ch in ALPHABET:
                    sub = prefix + ch
                    if db.prefix_status(sub) != "done" and sub not in queued:
                        queue.append(sub)
                        queued.add(sub)

            processed += 1
            now = time.time()
            if now - last_report >= 2.0:
                pc = db.prefixes_status_counts()
                counts = db.counts()
                # Queue size is "work remaining" for current run; "pending" is persisted.
                msg = (
                    f"[collect-pids] processed={processed} queue={len(queue)} "
                    f"prefixes(done={pc['done']}, error={pc['error']}) parts={counts['parts']}"
                )
                print(msg, file=sys.stderr)
                last_report = now

            time.sleep(sleep_s)
        except Exception as e:
            db.mark_prefix(prefix, "error", last_error=str(e))
            # Backoff a bit; keep going.
            time.sleep(max(1.0, sleep_s) * 3)


_DATE_RE = re.compile(
    r"(?P<from_m>\d{2})\.(?P<from_y>\d{2})\s*~\s*(?:(?P<to_m>\d{2})\.(?P<to_y>\d{2}))?"
)


def parse_application_date(text: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Parse dates like:
      "09.17 ~05.20" or "10.18 ~" or "07.22 ~"
    Return (date_from, date_to) as 'YYYY-MM' strings when possible.
    """
    t = re.sub(r"\\s+", " ", (text or "").strip())
    m = _DATE_RE.search(t)
    if not m:
        return None, None
    fm, fy = m.group("from_m"), m.group("from_y")
    tm, ty = m.group("to_m"), m.group("to_y")
    # Assume 20xx for two-digit years.
    date_from = f"20{fy}-{fm}" if fy and fm else None
    date_to = f"20{ty}-{tm}" if ty and tm else None
    return date_from, date_to


def extract_fitments_from_product_html(html: str) -> List[Dict[str, Any]]:
    soup = BeautifulSoup(html, "lxml")
    block = soup.select_one(".search-res-application")
    if not block:
        return []

    out: List[Dict[str, Any]] = []
    for item in block.select(".spollers__item"):
        title_a = item.select_one(".spollers__title a")
        if not title_a:
            continue
        title = re.sub(r"\\s+", " ", title_a.get_text(" ", strip=True))
        # Expected: "MAKE » Model"
        make, model = None, None
        if "»" in title:
            make, model = [x.strip() or None for x in title.split("»", 1)]
        else:
            make = title.strip() or None

        for line_a in item.select(".search-res-application__lines a.search-res-application__line"):
            href = line_a.get("href") or ""
            href_abs = urljoin(BASE_URL, href)
            qs = parse_qs(urlparse(href_abs).query)

            body_vals = [v.strip() for v in qs.get("body[]", []) if v.strip()]
            eng_no_vals = [v.strip() for v in qs.get("eng_no[]", []) if v.strip()]
            engine_cc_vals = [v.strip() for v in qs.get("engine[]", []) if v.strip()]
            year_vals = [v.strip() for v in qs.get("year[]", []) if v.strip()]

            article = line_a.select_one(".search-res-application__article")
            raw_article = re.sub(r"\\s+", " ", article.get_text(" ", strip=True)) if article else ""
            date_el = line_a.select_one(".application-date")
            date_from, date_to = parse_application_date(date_el.get_text(" ", strip=True) if date_el else "")

            text_el = line_a.select_one(".search-res-application__text")
            raw_text = re.sub(r"\\s+", " ", text_el.get_text(" ", strip=True)) if text_el else ""

            out.append(
                {
                    "make": make,
                    "model": model,
                    # In site terms this is often a body/generation code (e.g., X470).
                    "body": ", ".join(body_vals) if body_vals else None,
                    "engine_no": ", ".join(eng_no_vals) if eng_no_vals else None,
                    "engine_cc": ", ".join(engine_cc_vals) if engine_cc_vals else None,
                    # If year[] exists, prefer that as notes; otherwise use parsed dates.
                    "date_from": date_from,
                    "date_to": date_to,
                    "notes": ", ".join(year_vals) if year_vals else (raw_text or None),
                    "href": href_abs,
                    "raw_article": raw_article or None,
                    "raw_text": raw_text or None,
                }
            )

    return out


def fetch_product_page(client: HttpClient, pid: int) -> str:
    url = PRODUCT_URL_TMPL.format(pid=pid)
    raw = client.get_bytes(url, accept="text/html,application/xhtml+xml;q=0.9,*/*;q=0.8")
    return raw.decode("utf-8", "replace")


def scrape_pages(
    db: Db,
    client: HttpClient,
    *,
    html_cache_dir: Path,
    sleep_s: float,
    limit: Optional[int],
) -> None:
    ensure_dir(html_cache_dir)
    pids = db.list_part_pids()
    if limit is not None:
        pids = pids[:limit]

    total = len(pids)
    done_before = db.pages_status_counts().get("done", 0)
    for pid in tqdm(pids, desc="scrape-pages"):
        st = db.page_status(pid)
        if st == "done":
            continue

        try:
            html_path = html_cache_dir / f"{pid}.html"
            if html_path.exists():
                html = html_path.read_text(encoding="utf-8", errors="replace")
            else:
                html = fetch_product_page(client, pid)
                html_path.write_text(html, encoding="utf-8")

            fitments = extract_fitments_from_product_html(html)

            # Replace existing fitments for pid (idempotent).
            db.clear_fitments_for_pid(pid)
            if fitments:
                db.insert_fitments(pid, fitments)

            db.mark_page(pid, "done", http_status=200)
            time.sleep(sleep_s)
        except Exception as e:
            db.mark_page(pid, "error", last_error=str(e))
            time.sleep(max(1.0, sleep_s) * 3)

    # Summary
    ps = db.pages_status_counts()
    remaining = db.pages_remaining_estimate()
    counts = db.counts()
    print(
        f"[scrape-pages] pages(done={ps['done']}, error={ps['error']}), "
        f"remaining_estimate={remaining}, fitments={counts['fitments']}, total_parts={counts['parts']}",
        file=sys.stderr,
    )


def autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col = col_cells[0].column
        for c in col_cells:
            val = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col)].width = min(80, max(10, max_len + 2))


def export_xlsx(db: Db, out_path: Path) -> None:
    ensure_dir(out_path.parent)
    wb = Workbook()
    ws_parts = wb.active
    ws_parts.title = "Parts"
    ws_fit = wb.create_sheet("Fitment")

    parts_cols = [
        "pid",
        "part_no",
        "p_brand",
        "type_name",
        "house_part_no",
        "house_brand",
        "moved_id",
        "moved_part_no",
        "product_url",
        "first_seen_prefix",
        "created_at",
        "updated_at",
    ]
    ws_parts.append(parts_cols)
    for r in db.export_parts():
        rd = dict(r)
        row = []
        for c in parts_cols:
            if c == "product_url":
                row.append(PRODUCT_URL_TMPL.format(pid=rd.get("pid")))
            else:
                row.append(rd.get(c))
        ws_parts.append(row)

    fit_cols = [
        "pid",
        "make",
        "model",
        "generation_or_body",
        "engine_no",
        "engine_cc",
        "date_from",
        "date_to",
        "notes",
        "product_url",
        "href",
        "raw_article",
        "raw_text",
        "created_at",
    ]
    ws_fit.append(fit_cols)
    for r in db.export_fitments():
        rd = dict(r)
        row = []
        for c in fit_cols:
            if c == "generation_or_body":
                row.append(rd.get("body"))
            elif c == "product_url":
                row.append(PRODUCT_URL_TMPL.format(pid=rd.get("pid")))
            else:
                row.append(rd.get(c))
        ws_fit.append(row)

    autosize_columns(ws_parts)
    autosize_columns(ws_fit)
    wb.save(str(out_path))
    counts = db.counts()
    print(
        f"[export-xlsx] wrote={out_path} parts={counts['parts']} fitments={counts['fitments']}",
        file=sys.stderr,
    )


def print_status(db: Db) -> None:
    counts = db.counts()
    pc = db.prefixes_status_counts()
    ps = db.pages_status_counts()
    remaining = db.pages_remaining_estimate()
    print("DB:", db.path)
    print(f"parts: {counts['parts']}")
    print(f"fitments: {counts['fitments']}")
    print(f"prefixes: total={counts['prefixes']} done={pc['done']} error={pc['error']}")
    print(f"pages: total={counts['pages']} done={ps['done']} error={ps['error']}")
    print(f"pages_remaining_estimate: {remaining}")


def main(argv: Optional[Sequence[str]] = None) -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--db", default="out/cache/azfilter.sqlite3", help="SQLite cache/checkpoint path")
    p.add_argument("--html-cache", default="out/cache/html", help="Directory to cache product HTML pages")
    p.add_argument("--out", default="out/azfilter_parts.xlsx", help="Output XLSX path")
    p.add_argument("--sleep", type=float, default=0.15, help="Sleep between requests (seconds)")
    p.add_argument("--insecure", action="store_true", help="Disable TLS cert verification")
    p.add_argument("--max-prefix-len", type=int, default=10, help="Max prefix length for enumeration")

    sub = p.add_subparsers(dest="cmd", required=True)
    sub.add_parser("collect-pids")
    sp_scrape = sub.add_parser("scrape-pages")
    sp_scrape.add_argument("--limit", type=int, default=None, help="Limit number of parts to scrape (debug)")
    sub.add_parser("export-xlsx")
    sub.add_parser("status")
    sub.add_parser("run")

    args = p.parse_args(argv)

    db = Db(Path(args.db))
    client = HttpClient(insecure=bool(args.insecure))
    try:
        if args.cmd in ("collect-pids", "run"):
            collect_pids(
                db,
                client,
                sleep_s=float(args.sleep),
                max_prefix_len=int(args.max_prefix_len),
                start_prefixes=list(ALPHABET),
            )

        if args.cmd in ("scrape-pages", "run"):
            scrape_pages(
                db,
                client,
                html_cache_dir=Path(args.html_cache),
                sleep_s=float(args.sleep),
                limit=getattr(args, "limit", None),
            )

        if args.cmd in ("export-xlsx", "run"):
            export_xlsx(db, Path(args.out))

        if args.cmd == "status":
            print_status(db)

        return 0
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())

